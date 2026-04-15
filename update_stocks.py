"""
update_stocks.py
증권사에서 내보낸 엑셀 파일로 stocks.json 자동 업데이트

사용법: python update_stocks.py 잔고파일.xlsx
"""

import sys
import json
from openpyxl import load_workbook
from datetime import datetime

def update_from_excel(excel_path: str):
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # 기존 stocks.json 로드
    try:
        with open("stocks.json", encoding="utf-8") as f:
            config = json.load(f)
    except:
        config = {"korean": [], "us": []}

    # 기존 종목 딕셔너리 (이름 기준)
    existing = {s["name"]: s for s in config.get("korean", [])}

    # 엑셀에서 새 종목 파싱
    new_stocks = {}
    for row in rows[1:]:
        if not row[3] or row[3] == '종목명':
            continue
        name = str(row[3]).strip()
        new_stocks[name] = {
            "name":       name,
            "ticker":     existing.get(name, {}).get("ticker"),
            "buy_price":  round(float(row[16]), 2) if row[16] else None,
            "quantity":   int(row[13]) if row[13] else None,
            "active":     True,
        }

    # 기존에 있던 종목 중 엑셀에 없으면 active: false (매도된 종목)
    for name, stock in existing.items():
        if name not in new_stocks:
            stock["active"] = False
            new_stocks[name] = stock

    config["korean"] = list(new_stocks.values())

    with open("stocks.json", "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    active = sum(1 for s in config["korean"] if s.get("active"))
    print(f"완료: 활성 종목 {active}개 / 전체 {len(config['korean'])}개")
    print(f"업데이트 시각: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python update_stocks.py 잔고파일.xlsx")
        sys.exit(1)
    update_from_excel(sys.argv[1])
